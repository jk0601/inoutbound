"""
수불부 시트 생성 모듈

출고 = 입고 (기본 방침)
  → K열(출고수량)은 '=H{r}' 수식으로 삽입되므로
    필요 시 해당 셀에 숫자를 직접 입력하면 L(출고금액), M(재고)이 자동 재계산됨
"""
from collections import OrderedDict
from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 색상 상수 ─────────────────────────────────────────────────────────────────
_C = {
    'dark_blue':  '2F5496',   # 헤더 row1, 월계합계, 요약 헤더
    'light_blue': '8FAADC',   # 헤더 row2
    'green':      'E2EFDA',   # 분류 라벨행
    'gray':       'D6DCE4',   # 소계행, 요약합계행
    'info_label': 'D9E2F3',   # 정보 섹션 레이블 (관리부서 등)
    'highlight':  'FFF2CC',   # 전월이월·출고·재고 컬럼 강조
    'white':      'FFFFFF',
}

# 데이터행에서 배경 강조할 컬럼 번호 (F=6, J=10, K=11, L=12, M=13)
_HIGHLIGHT_COLS = {6, 10, 11, 12, 13}

_FONT_NAME = '맑은 고딕'
_MONEY_FMT = '#,##0'
_PCT_FMT   = '0.0%'


# ── 스타일 헬퍼 ───────────────────────────────────────────────────────────────

def _f(bold=False, size=9, color='000000'):
    return Font(name=_FONT_NAME, bold=bold, size=size, color=color)

def _fill(key):
    return PatternFill('solid', fgColor=_C[key])

def _al(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _side(style):
    return Side(border_style=style) if style else Side()

def _bd(left='thin', right='thin', top='thin', bottom='thin'):
    return Border(left=_side(left), right=_side(right),
                  top=_side(top),   bottom=_side(bottom))

def _apply(cell, font=None, fill=None, align=None, border=None, numfmt=None):
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    if numfmt: cell.number_format = numfmt


# ── 테이블 외곽 medium 테두리 적용 ────────────────────────────────────────────

def _outer_border(ws, min_row, max_row, min_col=1, max_col=14):
    """지정 범위의 외곽 셀에 medium 테두리를 적용한다. 내부는 thin 유지."""
    m = _side('medium')
    t = _side('thin')
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            c = ws.cell(row, col)
            c.border = Border(
                left   = m if col == min_col else t,
                right  = m if col == max_col else t,
                top    = m if row == min_row else t,
                bottom = m if row == max_row else t,
            )


# ── 셀 타입별 스타일 적용 ─────────────────────────────────────────────────────

def _h1(cell):
    """헤더 row1: 진한 파랑, 흰 볼드"""
    _apply(cell, _f(bold=True, size=9, color=_C['white']),
           _fill('dark_blue'), _al(), _bd())

def _h2(cell):
    """헤더 row2: 연한 파랑, 흰 볼드"""
    _apply(cell, _f(bold=True, size=9, color=_C['white']),
           _fill('light_blue'), _al(), _bd())

def _cat(cell, h='left'):
    """분류 라벨행: 연한 녹색, 볼드"""
    _apply(cell, _f(bold=True, size=9), _fill('green'), _al(h), _bd())

def _sub(cell, h='center', numfmt=None):
    """소계행: 회색, 볼드"""
    _apply(cell, _f(bold=True, size=9), _fill('gray'), _al(h), _bd(), numfmt)

def _tot(cell, h='center', numfmt=None):
    """월계합계행: 진한 파랑, 흰 볼드"""
    _apply(cell, _f(bold=True, size=10, color=_C['white']),
           _fill('dark_blue'), _al(h), _bd(), numfmt)

def _data(cell, h='center', numfmt=None, highlight=False):
    """일반 데이터행"""
    left = 'medium' if cell.column == 1 else 'thin'
    fill = _fill('highlight') if highlight else None
    _apply(cell, _f(size=9), fill, _al(h), _bd(left=left), numfmt)

def _info_label(cell):
    """정보행 레이블 (관리부서, 관리기간, 작성일)"""
    _apply(cell, _f(bold=True, size=10), _fill('info_label'), _al(), _bd())

def _info_val(cell):
    """정보행 값 (식품부, 2026년 3월 등) → 흰 배경"""
    _apply(cell, _f(size=10), _fill('white'), _al(), _bd())


# ── 그룹핑 ────────────────────────────────────────────────────────────────────

def _group(orders):
    groups = OrderedDict()
    for o in orders:
        groups.setdefault(o['분류'], []).append(o)
    return groups


# ── 수불부 시트 생성 ──────────────────────────────────────────────────────────

def generate_수불부(wb, orders: list, period: str, dept: str = '식품부'):
    if '수불부' in wb.sheetnames:
        del wb['수불부']
    ws = wb.create_sheet('수불부')

    groups = _group(orders)
    cats   = list(groups.keys())

    # ── 컬럼 너비 ────────────────────────────────────────────────────
    for col, w in [('A', 8.5), ('B', 13.5), ('C', 25.0), ('D', 19.0),
                   ('E', 6.7),  ('F', 7.5),  ('G', 12.0), ('H', 9.5),
                   ('I', 12.5), ('J', 12.0), ('K', 8.7),  ('L', 12.5),
                   ('M', 7.5),  ('N', 16.0)]:
        ws.column_dimensions[col].width = w

    # ── Row 1: 제목 ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    ws.merge_cells('A1:N1')
    ws['A1'].value = '소 모 품  수 불 부'
    _apply(ws['A1'], _f(bold=True, size=18), None, _al(), _bd())

    ws.row_dimensions[2].height = 8

    # ── Row 3: 정보 ──────────────────────────────────────────────────
    # 레이블: info_label(파랑), 값: 흰 배경  /  J3:N3는 무 스타일
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 6
    ws.merge_cells('B3:C3')
    ws.merge_cells('E3:F3')
    ws.merge_cells('H3:I3')

    for addr, val, is_label in [
        ('A3', '관리부서', True),  ('B3', dept,           False),
        ('D3', '관리기간', True),  ('E3', period,         False),
        ('G3', '작성일',  True),   ('H3', datetime.now(), False),
    ]:
        c = ws[addr]
        c.value = val
        (_info_label if is_label else _info_val)(c)

    ws['H3'].number_format = 'YYYY-MM-DD'

    # 병합 꼬리 셀도 흰 배경 + 테두리 통일
    for addr in ('C3', 'F3', 'I3'):
        _apply(ws[addr], fill=_fill('white'), border=_bd())

    # ── Rows 5-6: 테이블 헤더 ────────────────────────────────────────
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 20

    for col in range(1, 15):
        _h1(ws.cell(5, col))
        _h2(ws.cell(6, col))

    for rng in ('A5:A6', 'B5:B6', 'C5:C6', 'D5:D6', 'E5:E6',
                'G5:I5', 'J5:L5', 'N5:N6'):
        ws.merge_cells(rng)

    for addr, val in [('A5', 'No.'),        ('B5', '분류'),
                      ('C5', '품목명'),      ('D5', '규격'),
                      ('E5', '단위'),        ('F5', '전월이월'),
                      ('G5', '입고 (구매)'), ('J5', '출고 (사용)'),
                      ('M5', '재고'),        ('N5', '비고')]:
        ws[addr].value = val

    for addr, val in [('F6', '수량'),     ('G6', '일자'),   ('H6', '수량'),
                      ('I6', '금액(원)'),  ('J6', '일자'),   ('K6', '수량'),
                      ('L6', '금액(원)'),  ('M6', '수량')]:
        ws[addr].value = val

    # ── 데이터 영역 ──────────────────────────────────────────────────
    cur = 7
    subtotal_rows = {}
    item_counts   = {}
    no = 1

    h_map = {
        1:'center', 2:'left',   3:'left',   4:'left',
        5:'center', 6:'center', 7:'center', 8:'center',
        9:'right', 10:'center', 11:'center', 12:'right',
        13:'center', 14:'left',
    }

    for cat in cats:
        items = groups[cat]
        ws.row_dimensions[cur].height = 20

        # 분류 라벨행
        for col in range(1, 15):
            c = ws.cell(cur, col)
            c.value = cat if col == 2 else None
            _cat(c, h='center' if col == 2 else 'left')
        cur += 1

        item_start = cur

        for item in items:
            r = cur
            ws.row_dimensions[r].height = 20

            row_vals = [
                no,             # A: 연번
                item['분류'],   # B
                item['품목명'], # C
                item['규격'],   # D
                item['단위'],   # E
                None,           # F: 전월이월 (빈 셀 = 0)
                item['일자'],   # G: 입고일자
                item['수량'],   # H: 입고수량
                item['금액'],   # I: 입고금액
                f'=G{r}',               # J: 출고일자
                f'=H{r}',               # K: 출고수량 ← 수정 시 숫자 직접 입력
                f'=I{r}/H{r}*K{r}',    # L: 출고금액
                f'=F{r}+H{r}-K{r}',    # M: 재고
                None,           # N: 비고
            ]

            for col, val in enumerate(row_vals, 1):
                c = ws.cell(r, col)
                c.value = val
                numfmt = _MONEY_FMT if col in (9, 12) else None
                _data(c, h=h_map[col], numfmt=numfmt,
                      highlight=(col in _HIGHLIGHT_COLS))

            no  += 1
            cur += 1

        item_end = cur - 1

        # 소계행
        r = cur
        ws.row_dimensions[r].height = 20
        sub_map = {
            3:  f'{cat} 소계',
            8:  f'=SUM(H{item_start}:H{item_end})',
            9:  f'=SUM(I{item_start}:I{item_end})',
            11: f'=SUM(K{item_start}:K{item_end})',
            12: f'=SUM(L{item_start}:L{item_end})',
            13: f'=SUM(M{item_start}:M{item_end})',
        }
        for col in range(1, 15):
            c = ws.cell(r, col)
            c.value = sub_map.get(col)
            h = 'center' if col == 3 else ('right' if col in (9, 12) else 'center')
            _sub(c, h=h, numfmt=_MONEY_FMT if col in (9, 12) else None)

        subtotal_rows[cat] = r
        item_counts[cat]   = len(items)
        cur += 1

    # ── 공백 + 월계합계 ──────────────────────────────────────────────
    ws.row_dimensions[cur].height = 8
    cur += 1

    monthly_total_row = cur
    ws.row_dimensions[cur].height = 25
    st = list(subtotal_rows.values())

    total_map = {
        3:  '★ 월계 합계',
        8:  '=' + '+'.join(f'H{s}' for s in st),
        9:  '=' + '+'.join(f'I{s}' for s in st),
        11: '=' + '+'.join(f'K{s}' for s in st),
        12: '=' + '+'.join(f'L{s}' for s in st),
        13: '=' + '+'.join(f'M{s}' for s in st),
    }
    for col in range(1, 15):
        c = ws.cell(cur, col)
        c.value = total_map.get(col)
        h = 'center' if col == 3 else ('right' if col in (9, 12) else 'center')
        _tot(c, h=h, numfmt=_MONEY_FMT if col in (9, 12) else None)
    cur += 1

    # ── 메인 테이블 외곽 medium 테두리 ──────────────────────────────
    _outer_border(ws, 5, monthly_total_row, 1, 14)

    # ── 공백 + 월간 분류별 입출고 현황 ──────────────────────────────
    ws.row_dimensions[cur].height = 8
    cur += 1

    ws.cell(cur, 1).value = '■ 월간 분류별 입출고 현황'
    _apply(ws.cell(cur, 1), _f(bold=True, size=11), None, _al('left'))
    cur += 1

    ws.row_dimensions[cur].height = 6
    cur += 1

    # 요약 헤더
    summary_header_row = cur
    for col, val in enumerate(['분류', '입고건수', '입고수량', '입고금액(원)',
                                '구성비(%)', '출고건수', '출고수량',
                                '출고금액(원)', '재고수량'], 1):
        c = ws.cell(cur, col)
        c.value = val
        _h1(c)
    cur += 1

    data_start    = cur
    summary_total = cur + len(cats)

    for cat in cats:
        r   = cur
        sr  = subtotal_rows[cat]
        cnt = item_counts[cat]
        ws.row_dimensions[r].height = 18

        row_vals = [
            cat, cnt,
            f'=H{sr}', f'=I{sr}',
            f'=D{r}/$D${summary_total}',
            f'=B{r}',
            f'=K{sr}', f'=L{sr}', f'=M{sr}',
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(r, col)
            c.value = val
            h = 'left' if col == 1 else ('right' if col in (4, 8) else 'center')
            numfmt = _MONEY_FMT if col in (4, 8) else (_PCT_FMT if col == 5 else None)
            _apply(c, _f(size=9), None, _al(h), _bd(), numfmt)
        cur += 1

    # 요약 합계행
    r  = cur
    de = cur - 1
    ws.row_dimensions[r].height = 18
    assert r == summary_total

    sum_vals = ['합계'] + [f'=SUM({chr(64+col)}{data_start}:{chr(64+col)}{de})'
                           for col in range(2, 10)]
    for col, val in enumerate(sum_vals, 1):
        c = ws.cell(r, col)
        c.value = val
        h = 'left' if col == 1 else ('right' if col in (4, 8) else 'center')
        numfmt = _MONEY_FMT if col in (4, 8) else (_PCT_FMT if col == 5 else None)
        _sub(c, h=h, numfmt=numfmt)
    cur += 1

    # 요약 테이블 외곽 medium 테두리 (A:I 범위)
    _outer_border(ws, summary_header_row, summary_total, 1, 9)

    # ── 공백 + 확인 및 결재 ──────────────────────────────────────────
    ws.row_dimensions[cur].height = 8
    cur += 1

    ws.cell(cur, 1).value = '■ 확인 및 결재'
    _apply(ws.cell(cur, 1), _f(bold=True, size=11), None, _al('left'))
    cur += 1

    ws.row_dimensions[cur].height = 8
    cur += 1

    # 서명 테이블: 3개 박스 (작성자 A:B / 검토자 C:E / 승인자 F:G)
    sig_label_row = cur
    ws.row_dimensions[cur].height = 22
    cur += 1

    sig_sign_row = cur
    ws.row_dimensions[cur].height = 30

    # 병합 및 스타일 적용
    for (label_rng, sign_rng, label_val) in [
        (f'A{sig_label_row}:B{sig_label_row}', f'A{sig_sign_row}:B{sig_sign_row}', '작성자'),
        (f'C{sig_label_row}:E{sig_label_row}', f'C{sig_sign_row}:E{sig_sign_row}', '검토자'),
        (f'F{sig_label_row}:G{sig_label_row}', f'F{sig_sign_row}:G{sig_sign_row}', '승인자'),
    ]:
        ws.merge_cells(label_rng)
        ws.merge_cells(sign_rng)

        label_cell = ws[label_rng.split(':')[0]]
        sign_cell  = ws[sign_rng.split(':')[0]]

        label_cell.value = label_val
        sign_cell.value  = '(   서명   )'

        _apply(label_cell, _f(bold=True, size=10), _fill('info_label'), _al(), _bd())
        _apply(sign_cell,  _f(size=10),             _fill('white'),      _al(), _bd())

    # 서명 테이블 외곽 medium 테두리 (A:G 범위)
    _outer_border(ws, sig_label_row, sig_sign_row, 1, 7)
